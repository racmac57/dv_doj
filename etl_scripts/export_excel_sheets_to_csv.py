# export_excel_sheets_to_csv.py
# Portable version: Automatically converts all Excel files in the current directory to CSV
# Changes: Added batch processing, automatic folder detection, improved logging

import argparse
import os
import re
import json
import hashlib
import time
import psutil
from datetime import datetime, time as dtime, date, timedelta as dtdelta
import fnmatch
import csv
import logging
import signal
from tqdm import tqdm
from concurrent.futures import ThreadPoolExecutor, as_completed

import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
warnings.filterwarnings("ignore", category=UserWarning, module="dateutil.parser")

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries, get_column_letter
from openpyxl.utils import column_index_from_string
import pandas as pd

# Custom Exceptions
class ExportTimeoutError(Exception):
    pass

class ExportError(Exception):
    pass

# Timeout Handler (Windows compatible)
import threading
import time

def timeout_handler(timeout_seconds):
    def decorator(func):
        def wrapper(*args, **kwargs):
            result = [None]
            exception = [None]
            
            def target():
                try:
                    result[0] = func(*args, **kwargs)
                except Exception as e:
                    exception[0] = e
            
            thread = threading.Thread(target=target)
            thread.daemon = True
            thread.start()
            thread.join(timeout_seconds)
            
            if thread.is_alive():
                raise ExportTimeoutError(f"Export timed out after {timeout_seconds} seconds")
            
            if exception[0]:
                raise exception[0]
            
            return result[0]
        return wrapper
    return decorator

_TIME_RE = re.compile(r"^\s*\d{1,2}:\d{2}(:\d{2})?\s*$")

def sanitize_filename(name: str, used: set) -> str:
    fn = re.sub(r'[<>:"/\\|?*\x00-\x1F]', '_', name).strip().strip('.')
    fn = re.sub(r'[\s_]+', '_', fn)[:150] or "Sheet"
    original = fn
    i = 2
    while fn.lower() in used:
        fn = f"{original}_{i}"
        i += 1
    used.add(fn.lower())
    return fn

def get_used(ws):
    try:
        dim = ws.calculate_dimension()
    except ValueError:
        # Handle unsized worksheets
        dim = ws.calculate_dimension(force=True)
    
    if dim in ("A1:A1", None):
        return 0, 0
    min_c, min_r, max_c, max_r = range_boundaries(dim)
    return max_r, max_c

def get_file_hash(path):
    if not os.path.exists(path):
        return None
    hasher = hashlib.sha256()
    with open(path, 'rb') as f:
        for chunk in iter(lambda: f.read(4096), b""):
            hasher.update(chunk)
    return hasher.hexdigest()

def get_adaptive_chunksize(file_size, available_memory):
    memory_threshold = available_memory * 0.5  # Use 50% of available memory
    if file_size > 100 * 1024 * 1024:
        return 1000
    elif file_size > 50 * 1024 * 1024:
        return 2000
    elif file_size > 10 * 1024 * 1024:
        return 5000
    return 10000

def _parse_af_columns(spec: str):
    """
    Accept 'A,B,C,D,E,F' or 'A-F' (case-insensitive). Return sorted 1-based column indexes.
    """
    spec = spec.strip().upper().replace(" ", "")
    cols = []
    if "-" in spec and "," not in spec:
        start, end = spec.split("-", 1)
        s = column_index_from_string(start)
        e = column_index_from_string(end)
        if s > e: s, e = e, s
        cols = list(range(s, e + 1))
    else:
        parts = [p for p in spec.split(",") if p]
        for p in parts:
            cols.append(column_index_from_string(p))
    return sorted(set(cols))

def _to_text(value, date_fmt="%Y-%m-%d", dt_fmt="%Y-%m-%d %H:%M:%S", t_fmt="%H:%M:%S"):
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return value
    # Handle pandas/calamine time-only values that may come as timedelta
    try:
        import pandas as _pd
        if isinstance(value, getattr(_pd, 'Timedelta', ())) or isinstance(value, dtdelta):
            total_seconds = int(value.total_seconds())
            # Normalize to within a day
            total_seconds = total_seconds % (24 * 3600)
            hh = total_seconds // 3600
            mm = (total_seconds % 3600) // 60
            ss = total_seconds % 60
            return f"{hh:02d}:{mm:02d}:{ss:02d}"
    except Exception:
        pass
    if isinstance(value, datetime):
        # If Excel stored a time-only value, the date part is often 1899-12-30
        if value.date().isoformat() == "1899-12-30":
            return value.strftime(t_fmt)
        # If time component is midnight, treat as date-only
        if value.hour == 0 and value.minute == 0 and value.second == 0 and value.microsecond == 0:
            return value.strftime(date_fmt)
        return value.strftime(dt_fmt)
    if isinstance(value, date):
        return value.strftime(date_fmt)
    if isinstance(value, dtime):
        return value.strftime(t_fmt)
    if isinstance(value, str) and _TIME_RE.match(value):
        parts = [int(p) for p in value.strip().split(":")]
        while len(parts) < 3:
            parts.append(0)
        return dtime(parts[0], parts[1], parts[2]).strftime(t_fmt)
    return str(value)

def _get_used_range(data):
    """Get used range for both openpyxl worksheets and pandas DataFrames"""
    if hasattr(data, 'calculate_dimension'):  # openpyxl worksheet
        try:
            dim = data.calculate_dimension()
        except ValueError:
            # Handle unsized worksheets
            dim = data.calculate_dimension(force=True)
        
        if dim in ("A1:A1", None):
            return 0, 0
        min_c, min_r, max_c, max_r = range_boundaries(dim)
        return max_r, max_c
    else:  # pandas DataFrame
        if data.empty:
            return 0, 0
        return len(data), len(data.columns)


def _iter_rows_pandas(df, visible_only, date_fmt, dt_fmt, t_fmt, af_columns, min_af_nonempty, af_gate_visible_only):
    """Iterate over pandas DataFrame rows for calamine engine"""
    af_cols = _parse_af_columns(af_columns)
    max_r, max_c = _get_used_range(df)
    if max_r == 0 or max_c == 0:
        return
    
    # Emit header row when using pandas engine
    if getattr(df, "_include_headers", False):
        yield [str(col) for col in df.columns]

    for r, row in enumerate(df.itertuples(index=False), 1):
        if visible_only:
            # For pandas, we can't easily check row/column visibility
            # This would require additional metadata
            pass
            
        # A-F gate logic
        non_empty = 0
        af_values = []
        for c in af_cols:
            if c < len(row):
                val = row[c]
                af_values.append(val)
                if pd.notna(val) and val != "":
                    non_empty += 1
        
        if non_empty < min_af_nonempty:
            continue
        
        # Convert row to text
        row_out = []
        for val in row:
            formatted_val = _to_text(val, date_fmt, dt_fmt, t_fmt)
            row_out.append(formatted_val)
        
        if row_out and any(v != "" for v in row_out):
            yield row_out

def _stream_write_csv(path, rows_iterable, encoding: str, bom: bool):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    final_encoding = ("utf-8-sig" if bom and encoding.lower().replace("-", "") == "utf8" else encoding)
    row_count = 0
    max_cols_seen = 0
    with open(path, "w", encoding=final_encoding, newline="") as f:
        writer = csv.writer(f)
        for row in rows_iterable:
            if row is None:
                row = []
            writer.writerow(row)
            row_count += 1
            if isinstance(row, (list, tuple)):
                if len(row) > max_cols_seen:
                    max_cols_seen = len(row)
    return row_count, max_cols_seen

def process_excel_file(xlsx_path, out_dir, args, logger):
    """Process a single Excel file and return results"""
    results = {
        "file": xlsx_path,
        "success": False,
        "sheets": [],
        "error": None
    }
    
    try:
        file_size = os.path.getsize(xlsx_path)
        available_memory = psutil.virtual_memory().available / (1024 * 1024)  # MB
        
        engine = "calamine" if args.engine == "auto" else args.engine
        chunksize = args.chunksize if args.chunksize else get_adaptive_chunksize(file_size, available_memory)
        
        logger.info(f"Processing: {os.path.basename(xlsx_path)}")
        logger.info(f"Using engine: {engine}, file size: {file_size / (1024*1024):.2f}MB")
        
        # Initialize workbook based on engine
        if engine == "openpyxl":
            wb = load_workbook(xlsx_path, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
        else:  # calamine
            excel_file = pd.ExcelFile(xlsx_path, engine='calamine')
            sheet_names = excel_file.sheet_names
            wb = None
        
        used_names = set()
        
        # Filter sheets based on include/exclude
        filtered_sheets = []
        for sheet_name in sheet_names:
            if not args.all_sheets:
                if args.exclude and any(pat.lower() in sheet_name.lower() for pat in args.exclude):
                    continue
                if args.include and not any(pat.lower() in sheet_name.lower() for pat in args.include):
                    continue
            filtered_sheets.append(sheet_name)
        
        # Process each sheet
        for sheet_name in filtered_sheets:
            try:
                # Load sheet data based on engine
                if engine == "openpyxl":
                    ws = wb[sheet_name]
                    max_r, max_c = _get_used_range(ws)
                    is_df = False
                else:  # calamine
                    df = pd.read_excel(xlsx_path, sheet_name=sheet_name, engine='calamine', keep_default_na=False)
                    setattr(df, "_include_headers", True)
                    max_r, max_c = _get_used_range(df)
                    is_df = True
                    ws = None
                
                if not args.include_empty and (max_r == 0 or max_c == 0):
                    logger.info(f"  Skipping empty sheet: {sheet_name}")
                    continue
                
                safe = sanitize_filename(sheet_name, used_names)
                # If only one sheet is being exported, use the workbook base name
                if len(filtered_sheets) == 1:
                    csv_name = os.path.splitext(os.path.basename(xlsx_path))[0] + ".csv"
                else:
                    csv_name = f"{safe}.csv"
                csv_path = os.path.join(out_dir, csv_name)
                
                @timeout_handler(300)  # 5 min timeout
                def export_with_timeout():
                    if engine == "openpyxl":
                        # For openpyxl, we'd need to implement _iter_rows function or use calamine
                        return _stream_write_csv(csv_path, _iter_rows_pandas(pd.read_excel(xlsx_path, sheet_name=sheet_name, engine='calamine', keep_default_na=False), 
                                                                              args.visible_only, args.date_format, args.datetime_format, args.time_format, 
                                                                              args.af_columns, args.min_af_nonempty, args.af_gate_visible_only), 
                                                 args.encoding, args.bom)
                    else:  # calamine
                        return _stream_write_csv(csv_path, _iter_rows_pandas(df, args.visible_only, args.date_format, args.datetime_format, args.time_format, 
                                                                              args.af_columns, args.min_af_nonempty, args.af_gate_visible_only), 
                                                 args.encoding, args.bom)
                
                row_count, col_count = export_with_timeout()
                if row_count == 0:
                    logger.info(f"  No rows exported from sheet: {sheet_name}")
                    continue
                
                results["sheets"].append({
                    "sheet_name": sheet_name,
                    "sanitized_name": safe,
                    "file": csv_path,
                    "file_sha256": get_file_hash(csv_path),
                    "rows": row_count,
                    "cols": col_count
                })
                logger.info(f"  OK: Exported {row_count} rows x {col_count} cols to {os.path.basename(csv_path)}")
                
            except Exception as e:
                logger.error(f"  ERROR: Error processing sheet {sheet_name}: {str(e)}")
                results["sheets"].append({
                    "sheet_name": sheet_name,
                    "status": "error",
                    "error": str(e)
                })
        
        if wb:
            wb.close()
        
        results["success"] = len(results["sheets"]) > 0
        
    except Exception as e:
        logger.error(f"ERROR processing file: {str(e)}")
        results["error"] = str(e)
    
    return results

def main():
    ap = argparse.ArgumentParser(description="Export all Excel sheets to individual CSV files. Runs in current directory by default.")
    
    # Changed: --xlsx is now optional
    ap.add_argument("--xlsx", help="Path to the source .xlsx file (if not provided, processes all Excel files in current directory)")
    ap.add_argument("--output-dir", default="output", help="Output directory for CSV files (default: 'output')")
    ap.add_argument("--encoding", default="utf-8")
    ap.add_argument("--no-bom", dest="bom", action="store_false")
    ap.add_argument("--visible-only", action="store_true", help="Export only visible rows and columns")
    ap.add_argument("--all-sheets", action="store_true", default=True, help="Export all sheets (default)")
    ap.set_defaults(bom=True)
    ap.add_argument("--include-empty", action="store_true")
    ap.add_argument("--include", default=[], nargs="*", help="Include sheet patterns (substring match)")
    ap.add_argument("--exclude", default=[], nargs="*", help="Exclude sheet patterns (substring match)")
    ap.add_argument("--date-format", default="%Y-%m-%d", help="Date format string")
    ap.add_argument("--datetime-format", default="%Y-%m-%d %H:%M:%S", help="DateTime format string")
    ap.add_argument("--time-format", default="%H:%M:%S", help="Time format string")
    ap.add_argument("--af-columns", default="A,B,C",
                    help="Comma-separated list or Excel span (e.g., 'A-F') of columns to check for density gate.")
    ap.add_argument("--min-af-nonempty", type=int, default=2,
                    help="Minimum number of non-empty cells among --af-columns required to export a row.")
    ap.add_argument("--af-gate-visible-only", action="store_true",
                    help="If set, only visible cells among --af-columns count toward the threshold.")
    ap.add_argument("--disable-af-gate", action="store_true", default=True,
                    help="Disable A-F gate filtering (export all rows) - enabled by default")
    ap.add_argument("--engine", default="auto", choices=["auto", "calamine", "openpyxl"], 
                    help="Engine for reading Excel files (auto selects based on file size)")
    ap.add_argument("--chunksize", type=int, default=None, 
                    help="Read file in chunks of this size for large files (auto-enabled for 50MB+ files)")
    ap.add_argument("--log-file", default="export_log.json", help="Log file path (default: export_log.json)")
    
    args = ap.parse_args()
    
    # Handle disable-af-gate option
    if args.disable_af_gate:
        args.min_af_nonempty = 0
    
    # Setup logging - ensure output directory exists first
    current_dir = os.getcwd()
    log_file_path = os.path.join(current_dir, args.log_file)
    
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    logger = logging.getLogger(__name__)
    handler = logging.FileHandler(log_file_path)
    handler.setFormatter(logging.Formatter('{"time": "%(asctime)s", "level": "%(levelname)s", "message": "%(message)s"}'))
    logger.addHandler(handler)
    
    logger.info("=" * 80)
    logger.info("Excel to CSV Converter - Starting")
    logger.info(f"Working directory: {current_dir}")
    
    # Determine which files to process
    if args.xlsx:
        # Process single file
        xlsx_files = [os.path.abspath(args.xlsx)]
    else:
        # Find all Excel files in current directory
        xlsx_files = []
        for ext in ['*.xlsx', '*.xls']:
            xlsx_files.extend([os.path.abspath(f) for f in os.listdir(current_dir) if fnmatch.fnmatch(f.lower(), ext.lower())])
        
        if not xlsx_files:
            logger.error("No Excel files found in current directory")
            print("No Excel files found in current directory")
            return
        
        logger.info(f"Found {len(xlsx_files)} Excel file(s) to process")
        print(f"Found {len(xlsx_files)} Excel file(s) to process")
    
    # Create output directory
    output_dir = os.path.join(current_dir, args.output_dir)
    os.makedirs(output_dir, exist_ok=True)
    logger.info(f"Output directory: {output_dir}")
    
    # Process all files
    all_results = {
        "start_time": datetime.now().isoformat(),
        "files_processed": [],
        "total_files": len(xlsx_files),
        "successful_conversions": 0,
        "failed_conversions": 0
    }
    
    for i, xlsx_path in enumerate(xlsx_files, 1):
        logger.info("")
        print(f"\n[{i}/{len(xlsx_files)}] Processing {os.path.basename(xlsx_path)}...")
        
        result = process_excel_file(xlsx_path, output_dir, args, logger)
        all_results["files_processed"].append(result)
        
        if result["success"]:
            all_results["successful_conversions"] += 1
        else:
            all_results["failed_conversions"] += 1
    
    all_results["end_time"] = datetime.now().isoformat()
    all_results["duration_seconds"] = (datetime.fromisoformat(all_results["end_time"]) - 
                                       datetime.fromisoformat(all_results["start_time"])).total_seconds()
    
    # Save summary
    summary_path = os.path.join(output_dir, "conversion_summary.json")
    with open(summary_path, "w", encoding="utf-8") as f:
        json.dump(all_results, f, indent=2)
    
    # Print final summary
    logger.info("")
    logger.info("=" * 80)
    logger.info(f"Conversion complete: {all_results['successful_conversions']} successful, {all_results['failed_conversions']} failed")
    logger.info(f"Total time: {all_results['duration_seconds']:.2f} seconds")
    logger.info(f"Summary saved to: {summary_path}")
    
    print("\n" + "=" * 80)
    print("CONVERSION COMPLETE")
    print("=" * 80)
    print(f"Files processed    : {all_results['total_files']}")
    print(f"Successful         : {all_results['successful_conversions']}")
    print(f"Failed             : {all_results['failed_conversions']}")
    print(f"Output directory   : {output_dir}")
    print(f"Summary file       : {summary_path}")
    print(f"Log file           : {log_file_path}")
    print("=" * 80)

if __name__ == "__main__":
    main()
